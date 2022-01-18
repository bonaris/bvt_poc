import openpyxl as xl
from openpyxl import Workbook

MAX_TEST_DATA_COLUMNS = 1024
MAX_TEST_DATA_ROWS = 3000


@staticmethod
def get_test_data_record(filename, row_index):
    wb = xl.load_workbook(filename)
    sheet = wb['Search']



@staticmethod
def find_test_data(filename, tab_name, search_key):
    wb = xl.load_workbook(filename, read_only=True)

    test_data = None
    sheet = wb[tab_name]
    row_index = 1

    test_filed_list = []
    test_data_dictionary = {}

    total_columns = 0
    col = 1
    while col <= MAX_TEST_DATA_COLUMNS:
        value = sheet.cell(1, col).value
        if value is None:
            total_columns = col - 1
            break
        else:
            test_filed_list.append(sheet.cell(1, col).value)
            col += 1

    for row in sheet.rows:
        if sheet.cell(row_index, 1).value == search_key:
            break
        row_index += 1

    col = 1
    while col <= total_columns:
        test_data_dictionary.update({f'{test_filed_list[col-1]}': sheet.cell(row_index, col).value})
        col += 1
    return test_data_dictionary


@staticmethod
def find_all_records(filename, tab_name, search_key):
    wb = xl.load_workbook(filename, read_only=True)
    sheet = wb[tab_name]
    row_index = 1
    test_filed_list = []
    data_dictionary_list = []

    total_columns = 0
    col = 1
    while col <= MAX_TEST_DATA_COLUMNS:
        value = sheet.cell(1, col).value
        if value is None:
            total_columns = col - 1
            break
        else:
            test_filed_list.append(sheet.cell(1, col).value)
            col += 1

    for row in sheet.rows:
        if sheet.cell(row_index, 1).value == search_key:
            col = 1
            test_data_dictionary = {}
            while col <= total_columns:
                test_data_dictionary.update({f'{test_filed_list[col - 1]}': sheet.cell(row_index, col).value})
                col += 1
            data_dictionary_list.append(test_data_dictionary)
        row_index += 1
    return data_dictionary_list
